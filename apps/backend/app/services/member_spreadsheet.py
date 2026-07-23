from __future__ import annotations

from collections.abc import Iterable, Sequence
from dataclasses import dataclass, field
from datetime import date, datetime
from io import BytesIO
from typing import Any
from unicodedata import normalize
from uuid import UUID

from openpyxl import Workbook, load_workbook  # type: ignore[import-untyped]
from openpyxl.drawing.image import Image as ExcelImage  # type: ignore[import-untyped]
from PIL import Image as PilImage
from sqlalchemy import select
from sqlalchemy.exc import IntegrityError
from sqlalchemy.ext.asyncio import AsyncSession

from app.models import ElectoralMunicipality, ElectoralRegion, ElectoralState, Member

EXCEL_SHEET_NAME = "Datos"
# Order matches docs/Padron_Administrativo.xlsx (parse is by header name, not position).
EXCEL_HEADERS = (
    "Nro. CIV",
    "Nombre Completo",
    "Documento",
    "Correo electrónico",
    "Estatus",
    "Tipo",
    "Membresía",
    "Década",
    "Año de Graduación",
    "Sem",
    "Sexo",
    "Vivo",
    "Región",
    "Ubicación",
    "Municipio",
    "Seccional",
    "Título",
    "Mención",
    "Fecha Grado",
    "Foto",
)
OPTIONAL_HEADERS = frozenset({"Región", "Municipio", "Estado", "Ubicación", "Seccional", "Título"})
# Legacy header aliases → canonical EXCEL_HEADERS name
HEADER_ALIASES = {
    "código": "Nro. CIV",
    "codigo": "Nro. CIV",
    "nro. civ": "Nro. CIV",
    "nro civ": "Nro. CIV",
    "decada": "Década",
    "década": "Década",
    "año": "Año de Graduación",
    "ano": "Año de Graduación",
    "año de graduación": "Año de Graduación",
    "ano de graduacion": "Año de Graduación",
}
MAX_IMPORT_BYTES = 20 * 1024 * 1024


@dataclass(frozen=True)
class ParsedMember:
    row_number: int
    registry_code: str
    full_name: str
    dni: str
    email: str
    status: str
    member_type: str | None
    membership_months: int
    decade: int | None
    graduation_year: int | None
    semester: str | None
    sex: str | None
    alive: bool | None
    region: str | None
    section: str | None
    location: str | None
    ubicacion: str | None
    title: str | None
    mention: str | None
    graduation_date: date | None
    photo_filename: str | None


@dataclass(frozen=True)
class ImportRowError:
    row_number: int
    registry_code: str | None
    message: str


@dataclass
class MemberImportResult:
    rows_read: int = 0
    created: int = 0
    updated: int = 0
    failed: int = 0
    errors: list[ImportRowError] = field(default_factory=list)


class MemberSpreadsheetError(ValueError):
    """Raised when the workbook structure cannot be interpreted safely."""


def _header_key(value: Any) -> str:
    return normalize("NFKC", str(value or "").strip()).casefold()


def _text(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    text = str(value).strip()
    return text or None


def _required_text(value: Any, field_name: str) -> str:
    result = _text(value)
    if not result:
        raise ValueError(f"{field_name} is required")
    return result


def _optional_int(value: Any, field_name: str) -> int | None:
    text = _text(value)
    if text is None:
        return None
    try:
        return int(float(text))
    except ValueError as exc:
        raise ValueError(f"{field_name} must be an integer") from exc


def _optional_date(value: Any) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value).strip()
    for pattern in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(text, pattern).date()
        except ValueError:
            continue
    raise ValueError("Fecha Grado has an unsupported date format")


def _optional_alive(value: Any) -> bool | None:
    text = _text(value)
    if text is None:
        return None
    normalized = text.casefold()
    if normalized in {"1", "true", "t", "si", "sí", "vivo", "activo"}:
        return True
    if normalized in {"0", "false", "f", "no", "muerto", "inactivo"}:
        return False
    raise ValueError("Vivo must be 0/1 or a boolean value")


def _status(value: Any) -> str:
    text = _required_text(value, "Estatus").casefold()
    if text in {"activo", "active"}:
        return "ACTIVE"
    if text in {"inactivo", "inactive"}:
        return "INACTIVE"
    raise ValueError("Estatus must be Activo or Inactivo")


def _parse_row(row: tuple[Any, ...], columns: dict[str, int], row_number: int) -> ParsedMember:
    def value(header: str) -> Any:
        key = _header_key(header)
        if key not in columns:
            return None
        index = columns[key]
        return row[index] if index < len(row) else None

    return ParsedMember(
        row_number=row_number,
        registry_code=_required_text(value("Nro. CIV"), "Nro. CIV"),
        full_name=_required_text(value("Nombre Completo"), "Nombre Completo"),
        dni=_required_text(value("Documento"), "Documento"),
        email=_required_text(value("Correo electrónico"), "Correo electrónico").lower(),
        status=_status(value("Estatus")),
        member_type=_text(value("Tipo")),
        membership_months=_optional_int(value("Membresía"), "Membresía") or 0,
        decade=_optional_int(value("Década"), "Década"),
        graduation_year=_optional_int(value("Año de Graduación"), "Año de Graduación"),
        semester=_text(value("Sem")),
        sex=_text(value("Sexo")),
        alive=_optional_alive(value("Vivo")),
        region=_text(value("Región")),
        # Seccional is an organizational chapter in the reference file (often "Nacional").
        section=_text(value("Seccional")) or _text(value("Estado")),
        # Municipio drives municipality_id; Ubicación is the estado/state label in the reference file.
        location=_text(value("Municipio")),
        ubicacion=_text(value("Ubicación")),
        title=_text(value("Título")),
        mention=_text(value("Mención")),
        graduation_date=_optional_date(value("Fecha Grado")),
        photo_filename=_text(value("Foto")),
    )


def parse_member_workbook(content: bytes) -> tuple[list[ParsedMember], list[ImportRowError], int]:
    try:
        workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:
        raise MemberSpreadsheetError("The uploaded file is not a readable XLSX workbook") from exc

    if EXCEL_SHEET_NAME in workbook.sheetnames:
        sheet = workbook[EXCEL_SHEET_NAME]
    elif workbook.worksheets:
        sheet = workbook.worksheets[0]
    else:
        raise MemberSpreadsheetError("The workbook has no worksheets")

    rows = sheet.iter_rows(values_only=True)
    try:
        header_row = tuple(next(rows))
    except StopIteration as exc:
        raise MemberSpreadsheetError("The worksheet is empty") from exc

    columns = {
        _header_key(value): index for index, value in enumerate(header_row) if value is not None
    }
    # Normalize legacy aliases onto canonical header keys.
    for raw_key, index in list(columns.items()):
        canonical = HEADER_ALIASES.get(raw_key)
        if canonical is not None:
            columns.setdefault(_header_key(canonical), index)
    required_headers = tuple(h for h in EXCEL_HEADERS if h not in OPTIONAL_HEADERS)
    missing = [
        _header_key(header) for header in required_headers if _header_key(header) not in columns
    ]
    # Allow "Estado" as alias for the legacy "Seccional" column name.
    if _header_key("Seccional") in missing and _header_key("Estado") in columns:
        columns[_header_key("Seccional")] = columns[_header_key("Estado")]
        missing = [header for header in missing if header != _header_key("Seccional")]
    # Allow older workbooks without the newer optional columns.
    missing = [
        header
        for header in missing
        if header
        not in {
            _header_key("Municipio"),
            _header_key("Título"),
            _header_key("Ubicación"),
            _header_key("Seccional"),
        }
    ]
    if missing:
        raise MemberSpreadsheetError(f"Missing required columns: {', '.join(missing)}")

    parsed: list[ParsedMember] = []
    errors: list[ImportRowError] = []
    seen_codes: set[str] = set()
    seen_dnis: set[str] = set()
    rows_read = 0
    for row_number, raw_row in enumerate(rows, start=2):
        row = tuple(raw_row)
        if not any(value is not None and str(value).strip() for value in row):
            continue
        rows_read += 1
        code: str | None = None
        try:
            code = _text(row[columns[_header_key("Nro. CIV")]])
            if code and code in seen_codes:
                raise ValueError("Nro. CIV is duplicated in the workbook")
            member = _parse_row(row, columns, row_number)
            if member.dni in seen_dnis:
                raise ValueError(f"Documento '{member.dni}' is duplicated in the workbook")
            seen_codes.add(member.registry_code)
            seen_dnis.add(member.dni)
            parsed.append(member)
        except (TypeError, ValueError, KeyError) as exc:
            errors.append(ImportRowError(row_number, code, str(exc)))

    return parsed, errors, rows_read


def _member_values(
    member: ParsedMember,
    *,
    region_id: UUID | None = None,
    state_id: UUID | None = None,
    municipality_id: UUID | None = None,
    region_label: str | None = None,
    state_label: str | None = None,
    municipality_label: str | None = None,
) -> dict[str, Any]:
    # Persist estado text in `section` when Ubicación/territory resolves, so the UI
    # Estado fallback matches the reference workbook semantics.
    section_text = state_label or member.ubicacion or member.section
    location_text = municipality_label or member.location or member.ubicacion
    return {
        "registry_code": member.registry_code,
        "full_name": member.full_name,
        "dni": member.dni,
        "email": member.email,
        "status": member.status,
        "member_type": member.member_type,
        "membership_months": member.membership_months,
        "decade": member.decade,
        "graduation_year": member.graduation_year,
        "semester": member.semester,
        "sex": member.sex,
        "alive": member.alive,
        "region": member.region or region_label,
        "section": section_text,
        "location": location_text,
        "title": member.title,
        "mention": member.mention,
        "graduation_date": member.graduation_date,
        "photo_filename": member.photo_filename,
        "region_id": region_id,
        "state_id": state_id,
        "municipality_id": municipality_id,
    }


async def import_members(
    session: AsyncSession,
    organization_id: UUID,
    members: Sequence[ParsedMember],
    errors: list[ImportRowError],
    rows_read: int,
    dry_run: bool,
) -> MemberImportResult:
    regions = (
        await session.scalars(
            select(ElectoralRegion).where(ElectoralRegion.organization_id == organization_id)
        )
    ).all()
    states = (
        await session.scalars(
            select(ElectoralState).where(ElectoralState.organization_id == organization_id)
        )
    ).all()
    municipalities = (
        await session.scalars(
            select(ElectoralMunicipality).where(
                ElectoralMunicipality.organization_id == organization_id
            )
        )
    ).all()
    region_by_code = {r.code.strip().upper(): r.id for r in regions}
    region_by_name = {r.name.strip().upper(): r.id for r in regions}
    state_by_code = {s.code.strip().upper(): s for s in states}
    state_by_name = {s.name.strip().upper(): s for s in states}
    municipality_by_code = {m.code.strip().upper(): m for m in municipalities}
    municipality_by_name = {m.name.strip().upper(): m for m in municipalities}

    def resolve_region_id(label: str | None) -> UUID | None:
        if not label:
            return None
        key = label.strip().upper()
        return region_by_code.get(key) or region_by_name.get(key)

    def resolve_state(label: str | None) -> ElectoralState | None:
        if not label:
            return None
        key = label.strip().upper()
        return state_by_code.get(key) or state_by_name.get(key)

    def resolve_municipality(label: str | None) -> ElectoralMunicipality | None:
        if not label:
            return None
        key = label.strip().upper()
        return municipality_by_code.get(key) or municipality_by_name.get(key)

    result = MemberImportResult(rows_read=rows_read, errors=list(errors))
    for parsed in members:
        existing = await session.scalar(
            select(Member).where(
                Member.organization_id == organization_id,
                Member.registry_code == parsed.registry_code,
            )
        )
        matched_by = "registry_code" if existing is not None else None
        if existing is None:
            existing = await session.scalar(
                select(Member).where(
                    Member.organization_id == organization_id,
                    Member.dni == parsed.dni,
                )
            )
            if existing is not None:
                matched_by = "dni"
        if existing is None:
            existing = await session.scalar(
                select(Member).where(
                    Member.organization_id == organization_id,
                    Member.email == parsed.email,
                )
            )
            if existing is not None:
                matched_by = "email"

        if (
            existing is not None
            and matched_by != "registry_code"
            and existing.registry_code
            and existing.registry_code != parsed.registry_code
        ):
            result.errors.append(
                ImportRowError(
                    parsed.row_number,
                    parsed.registry_code,
                    (
                        f"Documento/correo ya pertenece al Nro. CIV {existing.registry_code}; "
                        "no se fusiona automáticamente"
                    ),
                )
            )
            continue

        try:
            created_this = False
            updated_this = False
            region_id = resolve_region_id(parsed.region)
            # Reference workbook: Ubicación holds the estado; Seccional is often "Nacional".
            state = resolve_state(parsed.ubicacion) or resolve_state(parsed.section)
            state_id = state.id if state else None
            state_label = state.name if state else None
            if state is not None and region_id is None:
                region_id = state.region_id
            municipality = resolve_municipality(parsed.location)
            municipality_id = municipality.id if municipality else None
            municipality_label = municipality.name if municipality else None
            if municipality is not None:
                state_id = municipality.state_id
                if state is None or state.id != municipality.state_id:
                    parent_state = next((s for s in states if s.id == municipality.state_id), None)
                    if parent_state is not None:
                        state = parent_state
                        state_label = parent_state.name
                        if region_id is None:
                            region_id = parent_state.region_id
            region_label = next((r.name for r in regions if r.id == region_id), None) if region_id else None
            values = _member_values(
                parsed,
                region_id=region_id,
                state_id=state_id,
                municipality_id=municipality_id,
                region_label=region_label,
                state_label=state_label or parsed.ubicacion,
                municipality_label=municipality_label,
            )
            async with session.begin_nested():
                if existing is None:
                    session.add(Member(organization_id=organization_id, **values))
                    result.created += 1
                    created_this = True
                else:
                    if values["photo_filename"] is None and existing.photo_data:
                        values.pop("photo_filename")
                    for field, value in values.items():
                        setattr(existing, field, value)
                    result.updated += 1
                    updated_this = True
                await session.flush()
        except IntegrityError:
            result.errors.append(
                ImportRowError(
                    parsed.row_number,
                    parsed.registry_code,
                    "Duplicate email, document or registry code in this organization",
                )
            )
            if created_this:
                result.created -= 1
            if updated_this:
                result.updated -= 1

    if dry_run:
        await session.rollback()
    else:
        await session.commit()
    result.failed = len(result.errors or [])
    return result


def _export_status(status: str) -> str:
    return "Activo" if status == "ACTIVE" else "Inactivo" if status == "INACTIVE" else status


def build_member_workbook(members: Iterable[Member]) -> BytesIO:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = EXCEL_SHEET_NAME
    sheet.append(EXCEL_HEADERS)
    sheet.freeze_panes = "A2"

    photo_buffers: list[BytesIO] = []
    for row_number, member in enumerate(members, start=2):
        sheet.append(
            [
                member.registry_code,
                member.full_name,
                member.dni,
                member.email,
                _export_status(member.status),
                member.member_type,
                member.membership_months,
                member.decade,
                member.graduation_year,
                member.semester,
                member.sex,
                1 if member.alive is True else 0 if member.alive is False else None,
                member.region,
                member.section,
                member.location,
                None,
                member.title,
                member.mention,
                member.graduation_date,
                member.photo_filename,
            ]
        )
        if member.photo_data:
            try:
                with PilImage.open(BytesIO(member.photo_data)) as source:
                    source.load()
                    image = source.convert("RGB")
                    buffer = BytesIO()
                    image.save(buffer, format="PNG")
                    buffer.seek(0)
                    photo_buffers.append(buffer)
                    excel_image = ExcelImage(buffer)
                    excel_image.width = 80
                    excel_image.height = 80
                    sheet.add_image(excel_image, f"T{row_number}")
                    sheet.row_dimensions[row_number].height = 60
            except Exception:
                # Keep the metadata row even if a legacy photo cannot be embedded.
                pass

    sheet.auto_filter.ref = sheet.dimensions
    for column in "ABCDEFGHIJKLMNOPQRST":
        sheet.column_dimensions[column].width = 18
    sheet.column_dimensions["B"].width = 32
    sheet.column_dimensions["D"].width = 30
    sheet.column_dimensions["T"].width = 24

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
